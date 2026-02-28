# -*- coding: utf-8 -*-
from odoo import models, fields, api
from odoo.exceptions import ValidationError
import base64
import json
import io
import openpyxl

class IcofrOnboardingWizard(models.TransientModel):
    _name = 'icofr.onboarding.wizard'
    _description = 'Master ICORF Onboarding Wizard'

    state = fields.Selection([
        ('setup', 'Setup'),
        ('fsli', 'FSLI Upload'),
        ('reconcile_entities', 'Entity Reconciliation'),
        ('gl', 'GL Upload'),
        ('rcm', 'RCM Mapping'),
        ('review', 'Review'),
        ('syncing', 'Syncing'),
        ('done', 'Done')
    ], string='Status', default='setup')

    session_id = fields.Many2one('icofr.onboarding.session', string='Active Session')
    
    # Setup fields
    company_id = fields.Many2one('res.company', string='Company', required=True, default=lambda self: self.env.company)
    fiscal_year = fields.Char(string='Fiscal Year', required=True, default=lambda self: str(fields.Date.today().year))
    sad_percent_override = fields.Float(string='SAD Percentage (%)', default=3.0, help='Juknis BUMN: 3% - 5%')

    # Upload fields
    fsli_file = fields.Binary(string='Template Laporan (FSLI)')
    fsli_filename = fields.Char(string='FSLI Filename')
    
    gl_file = fields.Binary(string='General Ledger')
    gl_filename = fields.Char(string='GL Filename')
    
    rcm_file = fields.Binary(string='Risk-Control Matrix (RCM)')
    rcm_filename = fields.Char(string='RCM Filename')

    # Progress and Stats
    progress_percent = fields.Float(related='session_id.progress_percent', string='Sync Progress')
    total_staging_lines = fields.Integer(compute='_compute_staging_stats')
    error_count = fields.Integer(compute='_compute_staging_stats')
    
    staging_error_ids = fields.One2many(
        'icofr.onboarding.staging.line', 
        compute='_compute_staging_errors', 
        string='Data Perlu Rekonsiliasi'
    )

    @api.depends('session_id', 'session_id.staging_line_ids')
    def _compute_staging_errors(self):
        for record in self:
            if record.session_id:
                record.staging_error_ids = record.session_id.staging_line_ids.filtered(lambda l: l.status == 'error')
            else:
                record.staging_error_ids = False

    @api.depends('session_id', 'session_id.staging_line_ids')
    def _compute_staging_stats(self):
        for record in self:
            if record.session_id:
                record.total_staging_lines = len(record.session_id.staging_line_ids)
                record.error_count = len(record.session_id.staging_line_ids.filtered(lambda l: l.status == 'error'))
            else:
                record.total_staging_lines = 0
                record.error_count = 0

    def action_start_session(self):
        self.ensure_one()
        session = self.env['icofr.onboarding.session'].search([
            ('company_id', '=', self.company_id.id),
            ('fiscal_year', '=', self.fiscal_year),
            ('state', '=', 'active')
        ], limit=1)
        
        if not session:
            session = self.env['icofr.onboarding.session'].create({
                'company_id': self.company_id.id,
                'fiscal_year': self.fiscal_year,
            })
        
        self.session_id = session
        self.state = 'fsli'
        return self._reopen_self()

    def action_previous(self):
        self.ensure_one()
        states = ['setup', 'fsli', 'reconcile_entities', 'gl', 'rcm', 'review']
        current_idx = states.index(self.state)
        if current_idx > 0:
            self.state = states[current_idx - 1]
        return self._reopen_self()

    def action_next(self):
        self.ensure_one()
        if self.state == 'fsli':
            return self.action_process_fsli()
        elif self.state == 'reconcile_entities':
            self.state = 'gl'
        elif self.state == 'gl':
            return self.action_process_gl()
        elif self.state == 'rcm':
            return self.action_process_rcm()
        elif self.state == 'review':
            self.state = 'syncing'
            return self.action_sync_start()
        return self._reopen_self()

    def _reopen_self(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_cancel(self):
        self.ensure_one()
        if self.session_id:
            self.session_id.action_cancel()
        return {'type': 'ir.actions.act_window_close'}

    # --- Excel Parser Engine ---

    def _validate_file_signature(self, file_content, file_type):
        try:
            file_data = base64.b64decode(file_content)
            wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
            sheet = wb.active
            header_row = next(sheet.iter_rows(min_row=1, max_row=1))
            headers = [str(cell.value).strip() for cell in header_row if cell.value is not None]
            
            signatures = {
                'fsli': ['Kode_BSPL', 'FSLI', 'Kode_Entity'],
                'gl': ['No_GL', 'GL_Balance', 'Kode_FSLI'],
                'rcm': ['ID Risiko', 'ID Kontrol']
            }
            required = signatures.get(file_type, [])
            missing = [h for h in required if h not in headers]
            if missing:
                raise ValidationError(f"Format file {file_type.upper()} tidak valid. Kolom wajib hilang: {', '.join(missing)}")
            return headers
        except Exception as e:
            if isinstance(e, ValidationError): raise e
            raise ValidationError(f"Gagal memvalidasi file: {str(e)}")

    def _parse_xlsx_to_staging(self, file_content, file_type):
        self.ensure_one()
        headers = self._validate_file_signature(file_content, file_type)
        file_data = base64.b64decode(file_content)
        wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
        sheet = wb.active
        
        staging_lines = []
        for row in sheet.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            if not any(v is not None for v in values): continue
            row_dict = {}
            for i, h in enumerate(headers):
                row_dict[h] = values[i] if i < len(values) else None
            
            staging_lines.append({
                'session_id': self.session_id.id,
                'type': file_type,
                'raw_data': json.dumps(row_dict, default=str),
                'status': 'draft'
            })
            if len(staging_lines) >= 500:
                self.env['icofr.onboarding.staging.line'].create(staging_lines)
                staging_lines = []
        if staging_lines:
            self.env['icofr.onboarding.staging.line'].create(staging_lines)

    def action_process_fsli(self):
        self.ensure_one()
        if not self.fsli_file:
            raise ValidationError("Silakan upload file Template Laporan (FSLI) terlebih dahulu.")
        self.session_id.staging_line_ids.filtered(lambda l: l.type == 'fsli').unlink()
        self._parse_xlsx_to_staging(self.fsli_file, 'fsli')
        
        staging_fsli = self.session_id.staging_line_ids.filtered(lambda l: l.type == 'fsli')
        existing_entity_codes = self.env['res.company'].search([]).mapped('entity_code')
        found_new = False
        for line in staging_fsli:
            data = line.get_data()
            e_code = str(data.get('Kode_Entity', '')).split('.')[0].strip()
            if e_code and e_code not in existing_entity_codes:
                found_new = True
                line.status = 'error'
                line.error_log = f"Entitas {e_code} tidak terdaftar."
        
        self.state = 'reconcile_entities' if found_new else 'gl'
        return self._reopen_self()

    def action_process_gl(self):
        self.ensure_one()
        if not self.gl_file:
            raise ValidationError("Silakan upload file General Ledger terlebih dahulu.")
        self.session_id.staging_line_ids.filtered(lambda l: l.type == 'gl').unlink()
        self._parse_xlsx_to_staging(self.gl_file, 'gl')
        self.state = 'rcm'
        return self._reopen_self()

    def action_process_rcm(self):
        self.ensure_one()
        if not self.rcm_file:
            raise ValidationError("Silakan upload file RCM terlebih dahulu.")
        self.session_id.staging_line_ids.filtered(lambda l: l.type == 'rcm').unlink()
        self._parse_xlsx_to_staging(self.rcm_file, 'rcm')
        self.state = 'review'
        return self._reopen_self()

    # --- Final Sync Engine (Task 7 & 8) ---

    def action_sync_start(self):
        """Final Step: Move data from staging to production models"""
        self.ensure_one()
        
        # 1. Create or Update Materiality Header
        mat_model = self.env['icofr.materiality']
        materiality = mat_model.search([
            ('company_id', '=', self.company_id.id),
            ('fiscal_year', '=', self.fiscal_year),
            ('active', '=', True)
        ], limit=1)
        
        if not materiality:
            materiality = mat_model.create({
                'name': f'Materialitas {self.fiscal_year} - {self.company_id.name}',
                'company_id': self.company_id.id,
                'fiscal_year': self.fiscal_year,
                'sad_percent': self.sad_percent_override,
            })
        
        # 2. Sync FSLI & GL (Account Mapping)
        staging_fsli = self.session_id.staging_line_ids.filtered(lambda l: l.type == 'fsli')
        mapping_model = self.env['icofr.account.mapping']
        
        for line in staging_fsli:
            data = line.get_data()
            bspl_code = str(data.get('Kode_BSPL', '')).strip()
            e_code = str(data.get('Kode_Entity', '')).split('.')[0].strip()
            
            if not bspl_code or bspl_code == '0.0': continue
            
            mapping_vals = {
                'fsl_item': bspl_code,
                'fsl_description': data.get('FSLI', ''),
                'entity_code': e_code,
                'materiality_id': materiality.id,
                'gl_account': 'ONBOARDING',
                'kategori': data.get('Kategori', ''),
                'sub_kategori': data.get('Sub_Kategori', ''),
            }
            
            # Find matching GL balance if uploaded
            gl_line = self.session_id.staging_line_ids.filtered(
                lambda l: l.type == 'gl' and l.get_data().get('Kode_FSLI') == bspl_code and l.get_data().get('Kode_Entity') == e_code
            )[:1]
            
            if gl_line:
                gl_data = gl_line.get_data()
                mapping_vals.update({
                    'gl_account': gl_data.get('No_GL', ''),
                    'gl_account_description': gl_data.get('GL_Desc', ''),
                    'account_balance': float(gl_data.get('GL_Balance', 0) or 0),
                    'status': 'mapped'
                })
            
            existing_mapping = mapping_model.search([
                ('fsl_item', '=', bspl_code),
                ('entity_code', '=', e_code),
                ('materiality_id', '=', materiality.id)
            ], limit=1)
            
            if existing_mapping:
                existing_mapping.write(mapping_vals)
            else:
                mapping_model.create(mapping_vals)

        # 3. Sync RCM (Risks & Controls)
        staging_rcm = self.session_id.staging_line_ids.filtered(lambda l: l.type == 'rcm')
        risk_model = self.env['icofr.risk']
        control_model = self.env['icofr.control']
        
        for line in staging_rcm:
            data = line.get_data()
            r_code = str(data.get('ID Risiko', '')).strip()
            c_code = str(data.get('ID Kontrol', '')).strip()
            
            if not r_code or not c_code: continue
            
            # Upsert Risk
            risk = risk_model.search([('code', '=', r_code)], limit=1)
            if not risk:
                risk = risk_model.create({
                    'code': r_code,
                    'name': data.get('Deskripsi Risiko', r_code),
                    'risk_category': 'operational',
                    'owner_id': self.env.user.id
                })
            
            # Upsert Control
            control = control_model.search([('code', '=', c_code)], limit=1)
            if not control:
                control = control_model.create({
                    'code': c_code,
                    'name': data.get('Deskripsi Kontrol', c_code),
                    'control_type': 'preventive',
                    'owner_id': self.env.user.id
                })
            
            # Relate Risk and Control
            if control not in risk.control_ids:
                risk.write({'control_ids': [(4, control.id)]})

        # Recalculate Materiality coverage
        materiality._compute_materiality_amounts()
        materiality._compute_scoping_coverage()
        
        self.session_id.action_complete()
        self.state = 'done'
        return self._reopen_self()
